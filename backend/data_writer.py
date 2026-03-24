# -*- coding: utf-8 -*-
"""
数据写入模块
负责将解析结果写入Excel和数据库
支持分Sheet存储不同器件类型
"""

import os
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .config import OUTPUT_DIR
from .db_manager import DatabaseManager
from .ai_processor import ExtractionResult, ExtractedParam

# 配置日志
logger = logging.getLogger(__name__)


class DataWriter:
    """
    数据写入器
    负责将提取结果写入Excel和数据库
    """
    
    # 样式定义
    HEADER_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    ERROR_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    SUCCESS_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def __init__(self, db_manager: DatabaseManager = None):
        """
        初始化数据写入器
        
        Args:
            db_manager: 数据库管理器实例
        """
        self.db_manager = db_manager or DatabaseManager()
        self.output_dir = OUTPUT_DIR
        self.output_dir.mkdir(exist_ok=True)
    
    def _get_param_columns(self) -> List[Dict[str, str]]:
        """
        获取参数列配置
        从数据库获取所有标准参数，生成Excel列
        
        Returns:
            列配置列表
        """
        # 固定列
        fixed_columns = [
            {'key': 'pdf_name', 'name': 'PDF文件名', 'width': 30},
            {'key': 'manufacturer', 'name': '厂家', 'width': 15},
            {'key': 'opn', 'name': 'OPN', 'width': 20},
            {'key': 'device_type', 'name': '器件类型', 'width': 15},
        ]
        
        # 从数据库获取参数列
        params = self.db_manager.get_all_standard_params()
        param_columns = []
        
        for param in params:
            param_columns.append({
                'key': param.param_name,
                'name': param.param_name,
                'width': 15,
                'unit': param.unit,
                'category': param.category
            })
        
        return fixed_columns + param_columns
    
    def _create_workbook_with_sheets(self) -> Workbook:
        """
        创建带有预定义Sheet的工作簿
        
        Returns:
            Workbook对象
        """
        wb = Workbook()
        
        # 删除默认Sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # 创建器件类型Sheet
        device_types = ['Si MOSFET', 'SiC MOSFET', 'IGBT', '汇总']
        
        for device_type in device_types:
            ws = wb.create_sheet(title=device_type)
            self._setup_sheet_headers(ws)
        
        return wb
    
    def _setup_sheet_headers(self, ws):
        """
        设置Sheet表头
        
        Args:
            ws: Worksheet对象
        """
        columns = self._get_param_columns()
        
        for col_idx, col_info in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_info['name'])
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self.BORDER
            
            # 设置列宽
            ws.column_dimensions[get_column_letter(col_idx)].width = col_info.get('width', 15)
        
        # 冻结首行
        ws.freeze_panes = 'A2'
    
    def write_to_excel(self, results: List[ExtractionResult], 
                       output_path: str = None,
                       append: bool = True) -> str:
        """
        将提取结果写入Excel
        
        Args:
            results: 提取结果列表
            output_path: 输出路径，默认自动生成
            append: 是否追加模式（追加到现有文件）
            
        Returns:
            输出文件路径
        """
        if output_path is None:
            output_path = self.output_dir / f"参数提取结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        else:
            output_path = Path(output_path)
        
        # 加载或创建工作簿
        if append and output_path.exists():
            wb = load_workbook(output_path)
        else:
            wb = self._create_workbook_with_sheets()
        
        columns = self._get_param_columns()
        column_keys = [col['key'] for col in columns]
        
        # 按器件类型分组写入
        for result in results:
            # 确定目标Sheet
            device_type = result.device_type or 'Si MOSFET'
            if device_type not in wb.sheetnames:
                ws = wb.create_sheet(title=device_type)
                self._setup_sheet_headers(ws)
            else:
                ws = wb[device_type]
            
            # 同时写入汇总Sheet
            ws_summary = wb['汇总'] if '汇总' in wb.sheetnames else None
            
            # 准备行数据
            row_data = {
                'pdf_name': result.pdf_name,
                'manufacturer': result.manufacturer,
                'opn': result.opn,
                'device_type': result.device_type
            }
            
            # 填充参数值
            for param in result.params:
                if param.standard_name in column_keys:
                    row_data[param.standard_name] = param.value
            
            # 写入数据行
            self._write_row(ws, row_data, column_keys, result.error)
            if ws_summary:
                self._write_row(ws_summary, row_data, column_keys, result.error)
        
        # 保存
        wb.save(output_path)
        logger.info(f"Excel文件已保存: {output_path}")
        
        return str(output_path)
    
    def _write_row(self, ws, row_data: Dict[str, Any], column_keys: List[str], 
                   error: str = None):
        """
        写入单行数据
        
        Args:
            ws: Worksheet对象
            row_data: 行数据字典
            column_keys: 列键列表
            error: 错误信息（如有）
        """
        row_idx = ws.max_row + 1
        
        for col_idx, key in enumerate(column_keys, 1):
            value = row_data.get(key, '')
            if value is None:
                value = ''
            
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
            cell.border = self.BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            # 错误行标红
            if error:
                cell.fill = self.ERROR_FILL
    
    def write_to_database(self, results: List[ExtractionResult], user_id: int = None) -> Dict[str, int]:
        """
        将提取结果写入数据库（支持用户隔离）
        
        Args:
            results: 提取结果列表
            user_id: 用户ID，用于数据隔离
            
        Returns:
            写入统计信息
        """
        stats = {
            'total': 0,
            'success': 0,
            'failed': 0
        }
        
        for result in results:
            # 先删除该PDF的旧结果（仅删除当前用户的）
            self.db_manager.delete_parse_results_by_pdf(result.pdf_name, user_id=user_id)
            
            if result.error:
                # 记录错误
                self.db_manager.add_parse_result(
                    pdf_name=result.pdf_name,
                    device_type=result.device_type,
                    manufacturer=result.manufacturer,
                    opn=result.opn,
                    is_success=False,
                    error_message=result.error,
                    user_id=user_id
                )
                self.db_manager.add_log(
                    log_type='ERROR',
                    content=f"解析失败: {result.error}",
                    pdf_name=result.pdf_name
                )
                stats['failed'] += 1
            else:
                # 写入参数
                for param in result.params:
                    self.db_manager.add_parse_result(
                        pdf_name=result.pdf_name,
                        device_type=result.device_type,
                        manufacturer=result.manufacturer,
                        opn=result.opn,
                        param_name=param.standard_name,
                        param_value=param.value,
                        test_condition=param.test_condition,
                        is_success=True,
                        user_id=user_id
                    )
                    stats['total'] += 1
                
                self.db_manager.add_log(
                    log_type='SUCCESS',
                    content=f"成功提取 {len(result.params)} 个参数",
                    pdf_name=result.pdf_name
                )
                stats['success'] += 1
        
        return stats
    
    def export_template(self, output_path: str = None) -> str:
        """
        导出空白Excel模板
        
        Args:
            output_path: 输出路径
            
        Returns:
            输出文件路径
        """
        if output_path is None:
            output_path = self.output_dir / "参数提取模板.xlsx"
        else:
            output_path = Path(output_path)
        
        wb = self._create_workbook_with_sheets()
        wb.save(output_path)
        
        logger.info(f"模板已导出: {output_path}")
        return str(output_path)
    
    def generate_report(self, results: List[ExtractionResult]) -> Dict[str, Any]:
        """
        生成解析报告
        
        Args:
            results: 提取结果列表
            
        Returns:
            报告数据
        """
        total_pdfs = len(results)
        success_pdfs = sum(1 for r in results if not r.error)
        failed_pdfs = total_pdfs - success_pdfs
        
        # 统计参数提取情况
        total_params = 0
        params_by_type = {}
        unrecognized = []
        
        for result in results:
            total_params += len(result.params)
            
            device_type = result.device_type or '未知'
            if device_type not in params_by_type:
                params_by_type[device_type] = 0
            params_by_type[device_type] += len(result.params)
            
            unrecognized.extend(result.unrecognized_params)
        
        # 计算完整率（假设每个PDF应提取约50个参数）
        expected_params = total_pdfs * 50
        completeness_rate = round(total_params / expected_params * 100, 2) if expected_params > 0 else 0
        
        report = {
            'summary': {
                'total_pdfs': total_pdfs,
                'success_pdfs': success_pdfs,
                'failed_pdfs': failed_pdfs,
                'success_rate': round(success_pdfs / total_pdfs * 100, 2) if total_pdfs > 0 else 0,
                'total_params': total_params,
                'completeness_rate': completeness_rate
            },
            'by_device_type': params_by_type,
            'unrecognized_params': list(set(unrecognized)),
            'failed_files': [r.pdf_name for r in results if r.error],
            'timestamp': datetime.now().isoformat()
        }
        
        # 完整率低于95%的建议
        if completeness_rate < 95:
            report['suggestions'] = [
                '建议检查参数库是否包含所有需要提取的参数',
                '检查PDF文件质量，确保表格可以被正确识别',
                '考虑添加更多参数变体名称以提高匹配率'
            ]
        
        return report
    
    def export_params_to_json(self, output_path: str = None) -> str:
        """
        导出参数库为JSON格式
        
        Args:
            output_path: 输出路径
            
        Returns:
            输出文件路径
        """
        import json
        
        if output_path is None:
            output_path = self.output_dir / "params_export.json"
        else:
            output_path = Path(output_path)
        
        params = self.db_manager.get_all_params_with_variants()
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(params, f, ensure_ascii=False, indent=2)
        
        logger.info(f"参数库已导出: {output_path}")
        return str(output_path)
    
    def import_params_from_json(self, json_path: str) -> int:
        """
        从JSON导入参数库
        
        Args:
            json_path: JSON文件路径
            
        Returns:
            导入的参数数量
        """
        import json
        
        with open(json_path, 'r', encoding='utf-8') as f:
            params = json.load(f)
        
        count = 0
        for param in params:
            result = self.db_manager.add_standard_param(
                param_name=param['param_name'],
                param_name_en=param.get('param_name_en'),
                param_type=param.get('param_type'),
                unit=param.get('unit'),
                category=param.get('category'),
                variants=param.get('variants', [])
            )
            if result:
                count += 1
        
        logger.info(f"成功导入 {count} 个参数")
        return count
    
    # ==================== 按条件生成表格 ====================
    
    # 未提取参数的橙色填充样式
    MISSING_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    MISSING_FONT = Font(color="D97706", italic=True)
    
    def generate_table_by_conditions(self, device_type: str, pdf_list: List[str],
                                     created_by: str = None, user_id: int = None) -> Dict[str, Any]:
        """
        按条件生成参数表格（支持用户隔离）
        
        核心逻辑：
        1. 根据器件类型+文件列表，从parse_results表中提取参数数据
        2. 关联standard_params表，获取该器件类型的所有标准参数（作为表格列）
        3. 按「每颗器件一行」的规则整理数据：
           - 行 = 选中的PDF文件
           - 列 = 该类型的所有标准参数
           - 单元格 = 参数值（含测试条件备注），未提取的参数标为「未提取」
        
        表格格式要求：
        - 表头：标准参数名（如「漏源击穿电压（V(BR)DSS）」）
        - 每行首列：PDF文件名
        - 参数值格式：「数值+单位（测试条件）」（如「100V（Tj=25℃）」）
        - 未提取的参数单元格标为橙色
        
        Args:
            device_type: 器件类型
            pdf_list: PDF文件列表
            created_by: 创建用户名
            user_id: 用户ID，用于数据隔离
            
        Returns:
            包含生成结果的字典：
            {
                'success': bool,
                'file_path': str,
                'table_name': str,
                'pdf_count': int,
                'param_count': int,
                'error': str (如有)
            }
        """
        import pandas as pd
        
        try:
            # 步骤1：获取表格数据（按用户过滤）
            logger.info(f"生成表格: 器件类型={device_type}, PDF列表={pdf_list}, user_id={user_id}")
            table_data = self.db_manager.get_params_for_table(device_type, pdf_list, user_id=user_id)
            
            logger.info(f"获取到 {len(table_data.get('data', []))} 行数据")
            
            if not table_data['data']:
                return {
                    'success': False,
                    'error': '未找到匹配的解析数据'
                }
            
            # 步骤2：生成文件名（按生成时间命名，避免覆盖）
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            safe_device_type = device_type.replace(' ', '_').replace('/', '_')
            table_name = f"{safe_device_type}_参数表_{timestamp}.xlsx"
            file_path = self.output_dir / table_name
            
            # 步骤3：创建Excel工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = device_type
            
            # 步骤4：写入表头
            headers = table_data['headers']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = self.HEADER_FILL
                cell.font = self.HEADER_FONT
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = self.BORDER
                
                # 设置列宽
                if col_idx == 1:
                    ws.column_dimensions[get_column_letter(col_idx)].width = 35  # PDF文件名列较宽
                elif col_idx <= 3:
                    ws.column_dimensions[get_column_letter(col_idx)].width = 18  # 型号/厂家列
                else:
                    ws.column_dimensions[get_column_letter(col_idx)].width = 15  # 参数列
            
            # 冻结首行
            ws.freeze_panes = 'A2'
            
            # 步骤5：写入数据行
            for row_idx, row_data in enumerate(table_data['data'], 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row_data.get(header, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else '')
                    cell.border = self.BORDER
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                    
                    # 未提取的参数标为橙色
                    if value == '-':
                        cell.fill = self.MISSING_FILL
                        cell.font = self.MISSING_FONT
            
            # 步骤6：保存Excel文件
            wb.save(file_path)
            logger.info(f"表格已生成: {file_path}")
            
            # 步骤7：保存表格记录到数据库
            record = self.db_manager.add_table_record(
                table_name=table_name,
                device_type=device_type,
                pdf_count=len(pdf_list),
                pdf_list=pdf_list,
                file_path=str(file_path),
                created_by=created_by
            )
            
            return {
                'success': True,
                'file_path': str(file_path),
                'table_name': table_name,
                'pdf_count': len(pdf_list),
                'param_count': table_data['param_count'],
                'record_id': record.id if record else None
            }
            
        except Exception as e:
            logger.error(f"生成表格失败: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def load_table_for_preview(self, file_path: str) -> Dict[str, Any]:
        """
        加载表格用于预览
        
        Args:
            file_path: 表格文件路径
            
        Returns:
            包含表格数据的字典
        """
        import pandas as pd
        
        try:
            if not os.path.exists(file_path):
                return {
                    'success': False,
                    'error': '文件不存在'
                }
            
            # 读取Excel文件
            df = pd.read_excel(file_path)
            
            # 转换为字典列表
            data = df.to_dict('records')
            headers = df.columns.tolist()
            
            # 标记未提取的单元格
            missing_cells = []
            for row_idx, row in enumerate(data):
                for col_idx, (key, value) in enumerate(row.items()):
                    if value == '-':
                        missing_cells.append({'row': row_idx, 'col': col_idx})
            
            return {
                'success': True,
                'data': data,
                'headers': headers,
                'row_count': len(data),
                'col_count': len(headers),
                'missing_cells': missing_cells
            }
            
        except Exception as e:
            logger.error(f"加载表格失败: {e}")
            return {
                'success': False,
                'error': str(e)
            }

